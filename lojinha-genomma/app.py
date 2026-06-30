"""
Lojinha Interna - Genomma Lab  |  Backend Python/Flask
"""
import os, json, threading, smtplib, logging, urllib.request, urllib.error, base64
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

from flask import Flask, request, jsonify, send_from_directory, Response
import openpyxl
import functools

app   = Flask(__name__, static_folder='public')
BASE  = Path(__file__).parent

# No Render, usa /tmp para arquivos mutáveis (stock, orders, uploads)
# Localmente, usa a própria pasta do projeto
IS_RENDER = os.getenv('RENDER', '') != ''
TMP       = Path('/tmp/lojinha') if IS_RENDER else BASE

DATA  = BASE / 'data'          # Excel fica sempre junto ao código
UPLOAD= TMP  / 'uploads'
STOCK = TMP  / 'stock.json'
ORDERS= TMP  / 'orders.json'
def _find_excel():
    """Retorna o .xlsx mais recente — busca em data/, raiz do projeto e subpastas"""
    files = sorted(BASE.rglob('*.xlsx'), key=lambda f: f.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f'Nenhum arquivo .xlsx encontrado em {BASE}')
    log.info(f'📂 Excel encontrado: {files[0].relative_to(BASE)}')
    return files[0]
LOCK  = threading.Lock()

UPLOAD.mkdir(parents=True, exist_ok=True)
(TMP).mkdir(parents=True, exist_ok=True)

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
SMTP_HOST  = 'smtp.gmail.com'
SMTP_PORT  = 587
SMTP_USER  = ''
SMTP_PASS  = ''
DEST_EMAIL = 'maycon.silva@contractor.genommalab.com'
PORT       = 3000
ADMIN_USER    = 'admin'
ADMIN_PASS    = '5827'
GITHUB_TOKEN  = ''
GITHUB_OWNER  = 'GanommaLab'
GITHUB_REPO   = 'lojinha-genomma'
GITHUB_BRANCH = 'main'
TEAMS_WEBHOOK_URL = 'https://defaultfd62b29cf0f442188ed60bccbb1b1b.67.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/8ea0e9a86e794cf499da0d84cb674621/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=V3ILIIj6HZlRCu1WfsTbzw3UNLpz8i_1ZdQxdLScpwk'

env_file = BASE / '.env'
if env_file.exists():
    for line in env_file.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip())
SMTP_HOST  = os.getenv('SMTP_HOST', SMTP_HOST)
SMTP_PORT  = int(os.getenv('SMTP_PORT', str(SMTP_PORT)))
SMTP_USER  = os.getenv('SMTP_USER', SMTP_USER)
SMTP_PASS  = os.getenv('SMTP_PASS', SMTP_PASS)
PORT       = int(os.getenv('PORT', str(PORT)))
ADMIN_PASS    = os.getenv('ADMIN_PASS',    ADMIN_PASS)
GITHUB_TOKEN  = os.getenv('GITHUB_TOKEN',  GITHUB_TOKEN)
GITHUB_OWNER  = os.getenv('GITHUB_OWNER',  GITHUB_OWNER)
GITHUB_REPO   = os.getenv('GITHUB_REPO',   GITHUB_REPO)
GITHUB_BRANCH = os.getenv('GITHUB_BRANCH', GITHUB_BRANCH)
TEAMS_WEBHOOK_URL = os.getenv('TEAMS_WEBHOOK_URL', TEAMS_WEBHOOK_URL)

# ── Backup GitHub ─────────────────────────────────────────────────────────────
def _gh_request(method, path, body=None):
    """Chama a API REST do GitHub. Retorna dict ou None em caso de erro."""
    if not GITHUB_TOKEN:
        return None
    url = f'https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{path}'
    headers = {
        'Authorization': f'token {GITHUB_TOKEN}',
        'Content-Type':  'application/json',
        'Accept':        'application/vnd.github.v3+json',
        'User-Agent':    'lojinha-genomma',
    }
    data = json.dumps(body).encode() if body else None
    req  = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None
        log.warning(f'⚠️  GitHub API {method} {path}: HTTP {e.code}')
        return None
    except Exception as e:
        log.warning(f'⚠️  GitHub API erro: {e}')
        return None

def _gh_backup(filename, content_str):
    """Salva arquivo na pasta backups/ do repositório GitHub."""
    if not GITHUB_TOKEN:
        return
    path     = f'backups/{filename}'
    encoded  = base64.b64encode(content_str.encode('utf-8')).decode()
    existing = _gh_request('GET', path)
    payload  = {
        'message': f'backup: {filename}',
        'content': encoded,
        'branch':  GITHUB_BRANCH,
    }
    if existing and 'sha' in existing:
        payload['sha'] = existing['sha']
    if _gh_request('PUT', path, payload):
        log.info(f'☁️  Backup GitHub OK → backups/{filename}')
    else:
        log.warning(f'⚠️  Backup GitHub falhou: backups/{filename}')

def _gh_restore(filename):
    """Lê arquivo da pasta backups/ do repositório GitHub."""
    if not GITHUB_TOKEN:
        return None
    result = _gh_request('GET', f'backups/{filename}')
    if result and 'content' in result:
        try:
            content = base64.b64decode(result['content']).decode('utf-8')
            log.info(f'☁️  Backup restaurado do GitHub: backups/{filename}')
            return content
        except Exception as e:
            log.warning(f'⚠️  Erro ao decodificar backup: {e}')
    return None

# ── Autenticação Admin ────────────────────────────────────────────────────────
def _check_auth(username, password):
    return username == ADMIN_USER and password == ADMIN_PASS

def require_admin(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not _check_auth(auth.username, auth.password):
            return Response(
                'Acesso restrito. Digite o usuário e senha.',
                401,
                {'WWW-Authenticate': 'Basic realm="Admin Lojinha Genomma"'}
            )
        return f(*args, **kwargs)
    return decorated

# ── Estoque ───────────────────────────────────────────────────────────────────
def load_from_excel():
    excel = _find_excel()
    log.info(f'📊 Lendo planilha Excel: {excel.name}')
    wb = openpyxl.load_workbook(str(excel), read_only=True, data_only=True)

    # ── Lê preços da aba "Tabela de Venda" (cabeçalho na linha 4, dados a partir da 5)
    prices = {}
    if 'Tabela de Venda' in wb.sheetnames:
        ws_tv = wb['Tabela de Venda']
        for row in ws_tv.iter_rows(min_row=5, values_only=True):
            if not row[0]: continue
            try:
                code  = str(int(float(row[0]))).strip()
                price = round(float(row[2]), 2) if row[2] is not None else 0.0
                prices[code] = price
            except: pass
        log.info(f'💲 {len(prices)} preços carregados da aba "Tabela de Venda".')

    # ── Lê estoque da aba "Estoque"
    ws = wb['Estoque']
    products = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        code = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] else ''
        try: qtd = float(row[13]) if row[13] is not None else 0
        except: qtd = 0
        if qtd > 0:
            if code not in products:
                products[code] = {'code': code, 'name': name, 'stock': 0, 'price': prices.get(code, 0.0)}
            products[code]['stock'] += qtd
    wb.close()
    for p in products.values(): p['stock'] = int(p['stock'])
    log.info(f'✅ {len(products)} produtos carregados.')
    return products

stock_data: dict = {}

def init_stock():
    global stock_data
    # 1) arquivo local (container ainda vivo)
    if STOCK.exists():
        try:
            stock_data = json.loads(STOCK.read_text('utf-8'))
            log.info(f'📦 Estoque local carregado ({len(stock_data)} produtos).')
            return
        except: pass
    # 2) backup no GitHub (após redeploy)
    backup = _gh_restore('stock.json')
    if backup:
        try:
            stock_data = json.loads(backup)
            STOCK.write_text(backup, 'utf-8')
            log.info(f'📦 Estoque restaurado do GitHub ({len(stock_data)} produtos).')
            return
        except: pass
    # 3) lê planilha Excel (primeira vez)
    stock_data = load_from_excel()
    content    = json.dumps(stock_data, ensure_ascii=False, indent=2)
    STOCK.write_text(content, 'utf-8')
    threading.Thread(target=_gh_backup, args=('stock.json', content), daemon=True).start()

def save_stock():
    content = json.dumps(stock_data, ensure_ascii=False, indent=2)
    STOCK.write_text(content, 'utf-8')
    threading.Thread(target=_gh_backup, args=('stock.json', content), daemon=True).start()

# ── Pedidos ───────────────────────────────────────────────────────────────────
def load_orders():
    # 1) arquivo local (container ainda vivo)
    if ORDERS.exists():
        try: return json.loads(ORDERS.read_text('utf-8'))
        except: pass
    # 2) backup no GitHub (após redeploy)
    backup = _gh_restore('orders.json')
    if backup:
        try:
            orders = json.loads(backup)
            ORDERS.write_text(backup, 'utf-8')
            log.info(f'📥 {len(orders)} pedidos restaurados do backup GitHub.')
            return orders
        except: pass
    return []

def write_orders(orders):
    content = json.dumps(orders, ensure_ascii=False, indent=2)
    ORDERS.write_text(content, 'utf-8')
    threading.Thread(target=_gh_backup, args=('orders.json', content), daemon=True).start()

# ── API: produtos ─────────────────────────────────────────────────────────────
@app.get('/api/products')
def api_products():
    with LOCK:
        lst = [p for p in stock_data.values() if p['stock'] > 0]
    lst.sort(key=lambda x: x['name'].lower())
    return jsonify(lst)

# ── API: pedido ───────────────────────────────────────────────────────────────
@app.post('/api/order')
def api_order():
    nome  = request.form.get('nome','').strip()
    email = request.form.get('email','').strip()
    tipo  = request.form.get('tipo_comprador','').strip()
    file  = request.files.get('comprovante')
    items_json = request.form.get('items_json','').strip()

    if not all([nome, email, tipo]):
        return jsonify({'error':'Preencha todos os campos obrigatórios.'}), 400

    # ── Pedido multi-item (novo formato) ──────────────────────────────────────
    if items_json:
        try:
            raw_items = json.loads(items_json)
            if not isinstance(raw_items, list) or not raw_items:
                return jsonify({'error':'Carrinho vazio.'}), 400
        except Exception:
            return jsonify({'error':'Dados de itens inválidos.'}), 400

        order_items = []
        total_valor = 0.0
        with LOCK:
            for item in raw_items:
                pcode = str(item.get('produto_code', '')).strip()
                try:
                    qty = int(float(item.get('quantidade', 0)))
                    if qty < 1: raise ValueError
                except Exception:
                    return jsonify({'error':f'Quantidade inválida para produto {pcode}.'}), 400
                product = stock_data.get(pcode)
                if not product:
                    return jsonify({'error':f'Produto não encontrado: {pcode}'}), 404
                if qty > product['stock']:
                    return jsonify({'error':f'Estoque insuficiente para "{product["name"]}". Disponível: {product["stock"]} un.'}), 400
                preco_unit = round(float(product.get('price', 0.0)), 2)
                valor_item = round(preco_unit * qty, 2)
                total_valor += valor_item
                order_items.append({
                    'produto_code': pcode,
                    'produto_name': product['name'],
                    'quantidade':   qty,
                    'preco_unit':   preco_unit,
                    'valor_total':  valor_item,
                })
                stock_data[pcode]['stock'] -= qty
            save_stock()

        attach_path, attach_name = None, None
        if file and file.filename:
            ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
            safe = ''.join(c if c.isalnum() or c in '._-' else '_' for c in file.filename)
            attach_name = f'{ts}_{safe}'
            attach_path = UPLOAD / attach_name
            file.save(str(attach_path))

        total_valor = round(total_valor, 2)
        order = {
            'id':          datetime.now().strftime('%Y%m%d%H%M%S%f')[:17],
            'data_hora':   datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            'nome':        nome,
            'email':       email,
            'tipo':        tipo,
            'items':       order_items,
            'valor_total': total_valor,
            'comprovante': attach_name or '',
            'status':      'pendente',
        }
        with LOCK:
            orders = load_orders()
            orders.insert(0, order)
            write_orders(orders)
        nomes = ', '.join(i['produto_name'] for i in order_items)
        log.info(f'📝 Pedido {order["id"]} — {nome} / {len(order_items)} itens: {nomes}')
        threading.Thread(target=_send_teams_notification, args=(order,), daemon=True).start()
        return jsonify({'success': True, 'message': 'Pedido finalizado!',
                        'itens': len(order_items), 'total': total_valor})

    # ── Pedido item único (backward compat) ───────────────────────────────────
    pcode = request.form.get('produto_code','').strip()
    qstr  = request.form.get('quantidade','0')
    if not all([pcode, qstr]):
        return jsonify({'error':'Preencha todos os campos obrigatórios.'}), 400
    try:
        qty = int(float(qstr))
        if qty < 1: raise ValueError
    except:
        return jsonify({'error':'Quantidade inválida.'}), 400

    with LOCK:
        product = stock_data.get(pcode)
        if not product: return jsonify({'error':'Produto não encontrado.'}), 404
        if qty > product['stock']:
            return jsonify({'error':f'Estoque insuficiente. Disponível: {product["stock"]} un.'}), 400
        stock_data[pcode]['stock'] -= qty
        save_stock()
        snap = dict(product)

    attach_path, attach_name = None, None
    if file and file.filename:
        ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe = ''.join(c if c.isalnum() or c in '._-' else '_' for c in file.filename)
        attach_name = f'{ts}_{safe}'
        attach_path = UPLOAD / attach_name
        file.save(str(attach_path))

    preco_unit  = round(float(snap.get('price', 0.0)), 2)
    valor_total = round(preco_unit * qty, 2)
    order = {
        'id':           datetime.now().strftime('%Y%m%d%H%M%S%f')[:17],
        'data_hora':    datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
        'nome':         nome,
        'email':        email,
        'tipo':         tipo,
        'produto_code': pcode,
        'produto_name': snap['name'],
        'quantidade':   qty,
        'preco_unit':   preco_unit,
        'valor_total':  valor_total,
        'comprovante':  attach_name or '',
        'status':       'pendente'
    }
    with LOCK:
        orders = load_orders()
        orders.insert(0, order)
        write_orders(orders)
    log.info(f'📝 Pedido {order["id"]} — {nome} / {snap["name"]} x{qty}')

    def try_email():
        try: _send_email(nome, email, tipo, snap, qty, attach_path, attach_name)
        except Exception as e: log.warning(f'⚠️ Email não enviado: {e}')
    threading.Thread(target=try_email, daemon=True).start()
    threading.Thread(target=_send_teams_notification, args=(order,), daemon=True).start()

    return jsonify({'success':True,'message':'Pedido finalizado!','produto':snap['name'],'quantidade':qty})

# ── API: listar pedidos ────────────────────────────────────────────────────────
@app.get('/api/orders')
@require_admin
def api_orders():
    return jsonify(load_orders())

# ── API: atualizar status ──────────────────────────────────────────────────────
@app.post('/api/orders/<order_id>/status')
@require_admin
def api_status(order_id):
    data       = request.get_json(force=True)
    new_status = data.get('status','pendente')
    with LOCK:
        orders = load_orders()
        found  = False
        for o in orders:
            if o.get('id') == order_id:
                o['status'] = new_status
                if new_status == 'entregue':
                    o['entregue_em'] = datetime.now().strftime('%d/%m/%Y %H:%M')
                else:
                    o.pop('entregue_em', None)
                found = True
                break
        if not found:
            return jsonify({'error':'Pedido não encontrado'}), 404
        write_orders(orders)
    return jsonify({'ok':True,'status':new_status})

# ── API: anexar nota fiscal ───────────────────────────────────────────────────
@app.post('/api/orders/<order_id>/nota_fiscal')
@require_admin
def api_nota_fiscal(order_id):
    file = request.files.get('nota_fiscal')
    if not file or not file.filename:
        return jsonify({'error': 'Nenhum arquivo enviado.'}), 400
    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe = ''.join(c if c.isalnum() or c in '._-' else '_' for c in file.filename)
    fname = f'nf_{order_id}_{ts}_{safe}'
    file.save(str(UPLOAD / fname))
    with LOCK:
        orders = load_orders()
        found  = False
        for o in orders:
            if o.get('id') == order_id:
                o['nota_fiscal'] = fname
                found = True
                break
        if not found:
            return jsonify({'error': 'Pedido não encontrado.'}), 404
        write_orders(orders)
    log.info(f'📄 Nota fiscal anexada ao pedido {order_id}: {fname}')
    return jsonify({'ok': True, 'nota_fiscal': fname})

# ── API: excluir pedido ───────────────────────────────────────────────────────
@app.delete('/api/orders/<order_id>')
@require_admin
def api_delete_order(order_id):
    with LOCK:
        orders = load_orders()
        found = None
        for o in orders:
            if o.get('id') == order_id:
                found = o
                break
        if not found:
            return jsonify({'error': 'Pedido não encontrado'}), 404
        orders.remove(found)
        write_orders(orders)
        # Restaurar estoque — suporta pedidos com múltiplos itens e item único
        stock_changed = False
        if 'items' in found:
            for item in found['items']:
                pcode = item.get('produto_code')
                qty   = int(item.get('quantidade', 0))
                if pcode and pcode in stock_data and qty > 0:
                    stock_data[pcode]['stock'] += qty
                    stock_changed = True
                    log.info(f'🔄 Estoque restaurado: +{qty} de {pcode}')
        else:
            pcode = found.get('produto_code')
            qty   = int(found.get('quantidade', 0))
            if pcode and pcode in stock_data and qty > 0:
                stock_data[pcode]['stock'] += qty
                stock_changed = True
                log.info(f'🔄 Estoque restaurado: +{qty} de {pcode}')
        if stock_changed:
            save_stock()
    return jsonify({'ok': True})

# ── API: inventário ───────────────────────────────────────────────────────────
@app.get('/api/inventario')
@require_admin
def api_inventario():
    with LOCK:
        orders = load_orders()
        # Calcula vendas por produto (suporta pedidos multi-item e item único)
        vendas = {}
        for o in orders:
            if 'items' in o:
                for item in o['items']:
                    pc  = item.get('produto_code')
                    qty = int(item.get('quantidade', 0))
                    if pc:
                        vendas[pc] = vendas.get(pc, 0) + qty
            else:
                pc  = o.get('produto_code')
                qty = int(o.get('quantidade', 0))
                if pc:
                    vendas[pc] = vendas.get(pc, 0) + qty
        # Monta relatório
        items = []
        for code, p in stock_data.items():
            sold    = vendas.get(code, 0)
            current = p['stock']
            initial = current + sold
            items.append({
                'code':    code,
                'name':    p['name'],
                'inicial': initial,
                'vendido': sold,
                'atual':   current,
            })
    items.sort(key=lambda x: x['name'].lower())
    return jsonify(items)

# ── Página Inventário ──────────────────────────────────────────────────────────
@app.get('/inventario')
@require_admin
def inventario():
    html = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Inventário — Lojinha Genomma</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',system-ui,sans-serif;background:#F3EDF9;min-height:100vh}
header{background:linear-gradient(135deg,#4A1B7A,#1A5FB4);padding:22px 32px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px}
header h1{color:white;font-size:1.4rem;font-weight:800}
header p{color:rgba(255,255,255,.75);font-size:.88rem}
.rbtn{background:rgba(255,255,255,.2);color:white;border:none;padding:7px 15px;border-radius:8px;cursor:pointer;font-size:.82rem;font-weight:600;text-decoration:none;display:inline-block}
.rbtn:hover{background:rgba(255,255,255,.32)}
.rbtn.green{background:#27AE60}.rbtn.green:hover{background:#219a52}
.rbtn.blue{background:#1A5FB4}.rbtn.blue:hover{background:#1550a0}
.container{max-width:1350px;margin:0 auto;padding:24px 18px}
.stats{display:flex;gap:13px;flex-wrap:wrap;margin-bottom:20px}
.stat{background:white;border-radius:13px;padding:15px 19px;flex:1;min-width:130px;box-shadow:0 2px 10px rgba(74,27,122,.09);border-left:4px solid #4A1B7A}
.stat.green{border-color:#27AE60}.stat.red{border-color:#E53935}.stat.orange{border-color:#E67E22}
.stat .n{font-size:1.7rem;font-weight:800;color:#4A1B7A}
.stat.green .n{color:#27AE60}.stat.red .n{color:#E53935}.stat.orange .n{color:#E67E22}
.stat .l{font-size:.75rem;color:#999;margin-top:1px}
.card{background:white;border-radius:15px;box-shadow:0 2px 14px rgba(74,27,122,.09);overflow:hidden;margin-bottom:20px}
.card-head{padding:15px 20px;border-bottom:1px solid #F0E8FB;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}
.card-head h2{font-size:.92rem;font-weight:700;color:#4A1B7A}
.actions{display:flex;gap:8px;flex-wrap:wrap}
table{width:100%;border-collapse:collapse}
thead th{background:#F8F3FF;padding:9px 12px;text-align:left;font-size:.68rem;font-weight:700;color:#7B3FAD;text-transform:uppercase;letter-spacing:.5px;white-space:nowrap}
thead th.num{text-align:center}
tbody tr{border-bottom:1px solid #FAF5FF;transition:background .13s}
tbody tr:hover{background:#FAF5FF}
tbody tr.diff-neg{background:#FFF8F8}
tbody tr.diff-pos{background:#F8FFF8}
td{padding:9px 12px;font-size:.84rem;vertical-align:middle}
td.num{text-align:center;font-weight:700}
td.code{font-family:monospace;font-size:.76rem;color:#999}
.chip{display:inline-block;padding:2px 9px;border-radius:20px;font-size:.72rem;font-weight:700}
.chip-sold{background:#FFF3E0;color:#E65100}
.chip-ok{background:#E8F5E9;color:#1B5E20}
.chip-warn{background:#FFEBEE;color:#B71C1C}
input.fisica{width:70px;text-align:center;border:2px solid #D1C4E9;border-radius:7px;padding:4px 6px;font-size:.85rem;font-weight:700;color:#4A1B7A;outline:none;transition:.15s}
input.fisica:focus{border-color:#7B3FAD;background:#FAF5FF}
.dif-pos{color:#1B5E20;font-weight:800}
.dif-neg{color:#B71C1C;font-weight:800}
.dif-zero{color:#999}
.info-box{background:#EDE7F6;border-left:4px solid #7B3FAD;padding:13px 16px;border-radius:8px;margin-bottom:18px;font-size:.85rem;color:#4A1B7A}
.info-box strong{display:block;margin-bottom:4px}
@media print{
  header .rbtn{display:none}
  .info-box{border:1px solid #ccc}
  input.fisica{border:1px solid #ccc}
}
</style>
</head>
<body>
<header>
  <div><h1>📊 Inventário</h1><p>Lojinha Interna — Genomma Lab</p></div>
  <div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center">
    <button class="rbtn green" onclick="exportCSV()">⬇️ Exportar CSV</button>
    <button class="rbtn blue"  onclick="window.print()">🖨️ Imprimir</button>
    <a href="/admin" class="rbtn">← Painel de Pedidos</a>
  </div>
</header>

<div class="container">
  <div class="info-box">
    <strong>📋 Como usar esta página</strong>
    Preencha a coluna <b>Contagem Física</b> com a quantidade que você contou fisicamente no estoque.
    A coluna <b>Diferença</b> calculará automaticamente (Física − Sistema).
    Valores negativos (em vermelho) indicam que o estoque físico está menor que o sistema.
  </div>

  <div class="stats">
    <div class="stat">       <div class="n" id="st-prod">—</div>  <div class="l">Produtos ativos</div></div>
    <div class="stat orange"><div class="n" id="st-ini">—</div>   <div class="l">Unidades iniciais</div></div>
    <div class="stat red">   <div class="n" id="st-vend">—</div>  <div class="l">Unidades vendidas</div></div>
    <div class="stat green"> <div class="n" id="st-atual">—</div> <div class="l">Em estoque (sistema)</div></div>
  </div>

  <div class="card">
    <div class="card-head">
      <h2>📦 Relatório de Estoque</h2>
      <span id="ref-label" style="font-size:.78rem;color:#aaa"></span>
    </div>
    <div style="overflow-x:auto">
      <table id="inv-table">
        <thead><tr>
          <th>#</th>
          <th>Código</th>
          <th>Produto</th>
          <th class="num">Est. Inicial</th>
          <th class="num">Vendido</th>
          <th class="num">Est. Sistema</th>
          <th class="num">Contagem Física</th>
          <th class="num">Diferença</th>
        </tr></thead>
        <tbody id="inv-body">
          <tr><td colspan="8" style="text-align:center;padding:40px;color:#aaa">Carregando...</td></tr>
        </tbody>
      </table>
    </div>
  </div>
</div>

<script>
let invData = [];

async function load() {
  try {
    const r = await fetch('/api/inventario');
    invData = await r.json();
    render();
  } catch(e) { console.error(e); }
}

function render() {
  const body = document.getElementById('inv-body');
  const now  = new Date().toLocaleDateString('pt-BR',{day:'2-digit',month:'2-digit',year:'numeric',hour:'2-digit',minute:'2-digit'});
  document.getElementById('ref-label').textContent = 'Gerado em: ' + now;

  let totIni=0, totVend=0, totAtual=0;
  invData.forEach(p => { totIni+=p.inicial; totVend+=p.vendido; totAtual+=p.atual; });
  document.getElementById('st-prod').textContent  = invData.length;
  document.getElementById('st-ini').textContent   = totIni;
  document.getElementById('st-vend').textContent  = totVend;
  document.getElementById('st-atual').textContent = totAtual;

  if (!invData.length) {
    body.innerHTML = "<tr><td colspan='8' style='text-align:center;padding:40px;color:#aaa'>Nenhum produto encontrado.</td></tr>";
    return;
  }

  body.innerHTML = invData.map((p, i) => {
    const hasSold = p.vendido > 0;
    const chip = hasSold
      ? `<span class='chip chip-sold'>-${p.vendido} vendido${p.vendido>1?'s':''}</span>`
      : `<span class='chip chip-ok'>sem saída</span>`;
    return `<tr id="row-${i}">
      <td style="color:#bbb;font-size:.76rem">${i+1}</td>
      <td class="code">${p.code}</td>
      <td style="font-weight:600;max-width:280px">${p.name}</td>
      <td class="num" style="color:#888">${p.inicial}</td>
      <td class="num">${chip}</td>
      <td class="num" style="color:#4A1B7A;font-size:1rem">${p.atual}</td>
      <td class="num"><input class="fisica" type="number" min="0" id="fis-${i}" placeholder="—" onchange="calcDif(${i})" oninput="calcDif(${i})"></td>
      <td class="num" id="dif-${i}" style="color:#ccc;font-size:1rem">—</td>
    </tr>`;
  }).join('');
}

function calcDif(i) {
  const val = document.getElementById('fis-'+i).value;
  const dif = document.getElementById('dif-'+i);
  const row = document.getElementById('row-'+i);
  if (val === '' || val === null) {
    dif.innerHTML = '<span class="dif-zero">—</span>';
    row.className = '';
    return;
  }
  const fisica  = parseInt(val, 10);
  const sistema = invData[i].atual;
  const diff    = fisica - sistema;
  if (diff > 0) {
    dif.innerHTML = `<span class="dif-pos">+${diff}</span>`;
    row.className = 'diff-pos';
  } else if (diff < 0) {
    dif.innerHTML = `<span class="dif-neg">${diff}</span>`;
    row.className = 'diff-neg';
  } else {
    dif.innerHTML = `<span class="dif-zero">✓ 0</span>`;
    row.className = '';
  }
}

function exportCSV() {
  const now = new Date().toLocaleDateString('pt-BR');
  let csv = '\\uFEFF'; // BOM para Excel reconhecer UTF-8
  csv += 'Código,Produto,Est. Inicial,Vendido,Est. Sistema,Contagem Física,Diferença\\n';
  invData.forEach((p, i) => {
    const fisEl = document.getElementById('fis-'+i);
    const fis   = fisEl && fisEl.value !== '' ? parseInt(fisEl.value, 10) : '';
    const dif   = fis !== '' ? fis - p.atual : '';
    const name  = '"' + p.name.replace(/"/g, '""') + '"';
    csv += `${p.code},${name},${p.inicial},${p.vendido},${p.atual},${fis},${dif}\\n`;
  });
  const blob = new Blob([csv], {type:'text/csv;charset=utf-8;'});
  const url  = URL.createObjectURL(blob);
  const a    = document.createElement('a');
  a.href = url;
  a.download = 'inventario_genomma_' + new Date().toISOString().slice(0,10) + '.csv';
  a.click();
  URL.revokeObjectURL(url);
}

load();
</script>
</body></html>"""
    return Response(html, mimetype='text/html')

# ── Relatório de vendas ───────────────────────────────────────────────────────
@app.get('/api/relatorio')
@require_admin
def api_relatorio():
    de_str  = request.args.get('de',  '')
    ate_str = request.args.get('ate', '')
    orders  = load_orders()

    def order_date(o):
        dh = o.get('data_hora', '')
        if not dh: return ''
        parts = dh.split(' ')[0].split('/')
        return f'{parts[code, p in stock_data.items():
            sold    = vendas.get(code, 0)
            current = p['stock']
            initial = current + sold
            items.append({
                'code':    code,
                'name':    p['name'],
                'inicial': initial,
                'vendido': sold,
                'atual':   current,
            })
    items.sort(key=lambda x: x['name'].lower())
    return jsonify(items)

# ── Página Inventário ─────────────────────────────────────────────────────────
@app.get('/inventario')
@require_admin
def inventario():
    html = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Inventário — Lojinha Genomma</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',system-ui,sans-serif;background:#F3EDF9;min-height:100vh}
header{background:linear-gradient(135deg,#4A1B7A,#1A5FB4);padding:22px 32px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px}
header h1{color:white;font-size:1.4rem;font-weight:800}
header p{color:rgba(255,255,255,.75);font-size:.88rem}
.rbtn{background:rgba(255,255,255,.2);color:white;border:none;padding:7px 15px;border-radius:8px;cursor:pointer;font-size:.82rem;font-weight:600;text-decoration:none;display:inline-block}
.rbtn:hover{background:rgba(255,255,255,.32)}
.rbtn.green{background:#27AE60}.rbtn.green:hover{background:#219a52}
.rbtn.blue{background:#1A5FB4}.rbtn.blue:hover{background:#1550a0}
.container{max-width:1350px;margin:0 auto;padding:24px 18px}
.stats{display:flex;gap:13px;flex-wrap:wrap;margin-bottom:20px}
.stat{background:white;border-radius:13px;padding:15px 19px;flex:1;min-width:130px;box-shadow:0 2px 10px rgba(74,27,122,.09);border-left:4px solid #4A1B7A}
.stat.green{border-color:#27AE60}.stat.red{border-color:#E53935}.stat.orange{border-color:#E67E22}
.stat .n{font-size:1.7rem;font-weight:800;color:#4A1B7A}
.stat.green .n{color:#27AE60}.stat.red .n{color:#E53935}.stat.orange .n{color:#E67E22}
.stat .l{font-size:.75rem;color:#999;margin-top:1px}
.card{background:white;border-radius:15px;box-shadow:0 2px 14px rgba(74,27,122,.09);overflow:hidden;margin-bottom:20px}
.card-head{padding:15px 20px;border-bottom:1px solid #F0E8FB;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}
.card-head h2{font-size:.92rem;font-weight:700;color:#4A1B7A}
.actions{display:flex;gap:8px;flex-wrap:wrap}
table{width:100%;border-collapse:collapse}
thead th{background:#F8F3FF;padding:9px 12px;text-align:left;font-size:.68rem;font-weight:700;color:#7B3FAD;text-transform:uppercase;letter-spacing:.5px;white-space:nowrap}
thead th.num{text-align:center}
tbody tr{border-bottom:1px solid #FAF5FF;transition:background .13s}
tbody tr:hover{background:#FAF5FF}
tbody tr.diff-neg{background:#FFF8F8}
tbody tr.diff-pos{background:#F8FFF8}
td{padding:9px 12px;font-size:.84rem;vertical-align:middle}
td.num{text-align:center;font-weight:700}
td.code{font-family:monospace;font-size:.76rem;color:#999}
.chip{display:inline-block;padding:2px 9px;border-radius:20px;font-size:.72rem;font-weight:700}
.chip-sold{background:#FFF3E0;color:#E65100}
.chip-ok{background:#E8F5E9;color:#1B5E20}
.chip-warn{background:#FFEBEE;color:#B71C1C}
input.fisica{width:70px;text-align:center;border:2px solid #D1C4E9;border-radius:7px;padding:4px 6px;font-size:.85rem;font-weight:700;color:#4A1B7A;outline:none;transition:.15s}
input.fisica:focus{border-color:#7B3FAD;background:#FAF5FF}
.dif-pos{color:#1B5E20;font-weight:800}
.dif-neg{color:#B71C1C;font-weight:800}
.dif-zero{color:#999}
.info-box{background:#EDE7F6;border-left:4px solid #7B3FAD;padding:13px 16px;border-radius:8px;margin-bottom:18px;font-size:.85rem;color:#4A1B7A}
.info-box strong{display:block;margin-bottom:4px}
@media print{
  header .rbtn{display:none}
  .info-box{border:1px solid #ccc}
  input.fisica{border:1px solid #ccc}
}
</style>
</head>
<body>
<header>
  <div><h1>📊 Inventário</h1><p>Lojinha Interna — Genomma Lab</p></div>
  <div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center">
    <button class="rbtn green" onclick="exportCSV()">⬇️ Exportar CSV</button>
    <button class="rbtn blue"  onclick="window.print()">🖨️ Imprimir</button>
    <a href="/admin" class="rbtn">← Painel de Pedidos</a>
  </div>
</header>

<div class="container">
  <div class="info-box">
    <strong>📋 Como usar esta página</strong>
    Preencha a coluna <b>Contagem Física</b> com a quantidade que você contou fisicamente no estoque.
    A coluna <b>Diferença</b> calculará automaticamente (Física − Sistema).
    Valores negativos (em vermelho) indicam que o estoque físico está menor que o sistema.
  </div>

  <div class="stats">
    <div class="stat">       <div class="n" id="st-prod">—</div>  <div class="l">Produtos ativos</div></div>
    <div class="stat orange"><div class="n" id="st-ini">—</div>   <div class="l">Unidades iniciais</div></div>
    <div class="stat red">   <div class="n" id="st-vend">—</div>  <div class="l">Unidades vendidas</div></div>
    <div class="stat green"> <div class="n" id="st-atual">—</div> <div class="l">Em estoque (sistema)</div></div>
  </div>

  <div class="card">
    <div class="card-head">
      <h2>📦 Relatório de Estoque</h2>
      <span id="ref-label" style="font-size:.78rem;color:#aaa"></span>
    </div>
    <div style="overflow-x:auto">
      <table id="inv-table">
        <thead><tr>
          <th>#</th>
          <th>Código</th>
          <th>Produto</th>
          <th class="num">Est. Inicial</th>
          <th class="num">Vendido</th>
          <th class="num">Est. Sistema</th>
          <th class="num">Contagem Física</th>
          <th class="num">Diferença</th>
        </tr></thead>
        <tbody id="inv-body">
          <tr><td colspan="8" style="text-align:center;padding:40px;color:#aaa">Carregando...</td></tr>
        </tbody>
      </table>
    </div>
  </div>
</div>

<script>
let invData = [];

async function load() {
  try {
    const r = await fetch('/api/inventario');
    invData = await r.json();
    render();
  } catch(e) { console.error(e); }
}

function render() {
  const body = document.getElementById('inv-body');
  const now  = new Date().toLocaleDateString('pt-BR',{day:'2-digit',month:'2-digit',year:'numeric',hour:'2-digit',minute:'2-digit'});
  document.getElementById('ref-label').textContent = 'Gerado em: ' + now;

  let totIni=0, totVend=0, totAtual=0;
  invData.forEach(p => { totIni+=p.inicial; totVend+=p.vendido; totAtual+=p.atual; });
  document.getElementById('st-prod').textContent  = invData.length;
  document.getElementById('st-ini').textContent   = totIni;
  document.getElementById('st-vend').textContent  = totVend;
  document.getElementById('st-atual').textContent = totAtual;

  if (!invData.length) {
    body.innerHTML = "<tr><td colspan='8' style='text-align:center;padding:40px;color:#aaa'>Nenhum produto encontrado.</td></tr>";
    return;
  }

  body.innerHTML = invData.map((p, i) => {
    const hasSold = p.vendido > 0;
    const chip = hasSold
      ? `<span class='chip chip-sold'>-${p.vendido} vendido${p.vendido>1?'s':''}</span>`
      : `<span class='chip chip-ok'>sem saída</span>`;
    return `<tr id="row-${i}">
      <td style="color:#bbb;font-size:.76rem">${i+1}</td>
      <td class="code">${p.code}</td>
      <td style="font-weight:600;max-width:280px">${p.name}</td>
      <td class="num" style="color:#888">${p.inicial}</td>
      <td class="num">${chip}</td>
      <td class="num" style="color:#4A1B7A;font-size:1rem">${p.atual}</td>
      <td class="num"><input class="fisica" type="number" min="0" id="fis-${i}" placeholder="—" onchange="calcDif(${i})" oninput="calcDif(${i})"></td>
      <td class="num" id="dif-${i}" style="color:#ccc;font-size:1rem">—</td>
    </tr>`;
  }).join('');
}

function calcDif(i) {
  const val = document.getElementById('fis-'+i).value;
  const dif = document.getElementById('dif-'+i);
  const row = document.getElementById('row-'+i);
  if (val === '' || val === null) {
    dif.innerHTML = '<span class="dif-zero">—</span>';
    row.className = '';
    return;
  }
  const fisica  = parseInt(val, 10);
  const sistema = invData[i].atual;
  const diff    = fisica - sistema;
  if (diff > 0) {
    dif.innerHTML = `<span class="dif-pos">+${diff}</span>`;
    row.className = 'diff-pos';
  } else if (diff < 0) {
    dif.innerHTML = `<span class="dif-neg">${diff}</span>`;
    row.className = 'diff-neg';
  } else {
    dif.innerHTML = `<span class="dif-zero">✓ 0</span>`;
    row.className = '';
  }
}

function exportCSV() {
  const now = new Date().toLocaleDateString('pt-BR');
  let csv = '\uFEFF'; // BOM para Excel reconhecer UTF-8
  csv += 'Código,Produto,Est. Inicial,Vendido,Est. Sistema,Contagem Física,Diferença\n';
  invData.forEach((p, i) => {
    const fisEl = document.getElementById('fis-'+i);
    const fis   = fisEl && fisEl.value !== '' ? parseInt(fisEl.value, 10) : '';
    const dif   = fis !== '' ? fis - p.atual : '';
    const name  = '"' + p.name.replace(/"/g, '""') + '"';
    csv += `${p.code},${name},${p.inicial},${p.vendido},${p.atual},${fis},${dif}\n`;
  });
  const blob = new Blob([csv], {type:'text/csv;charset=utf-8;'});
  const url  = URL.createObjectURL(blob);
  const a    = document.createElement('a');
  a.href = url;
  a.download = 'inventario_genomma_' + new Date().toISOString().slice(0,10) + '.csv';
  a.click();
  URL.revokeObjectURL(url);
}

load();
</script>
</body></html>"""
    return Response(html, mimetype='text/html')

# ── Relatório de vendas ─────────────────────────────────────────────────────────
@app.get('/api/relatorio')
@require_admin
def api_relatorio():
    de_str  = request.args.get('de',  '')
    ate_str = request.args.get('ate', '')
    orders  = load_orders()

    def order_date(o):
        dh = o.get('data_hora', '')
        if not dh: return ''
        parts = dh.split(' ')[0].split('/')
        return f'{parts[2]}-{parts[1]}-{parts[0]}' if len(parts) == 3 else ''

    filtered = [o for o in orders if
                (not de_str  or order_date(o) >= de_str) and
                (not ate_str or order_date(o) <= ate_str)]

    def order_qty(o):
        if o.get('items'):
            return sum(int(i.get('quantidade') or 0) for i in o['items'])
        return int(o.get('quantidade') or 0)

    total_pedidos = len(filtered)
    valor_total   = sum(float(o.get('valor_total') or 0) for o in filtered)
    ticket_medio  = valor_total / total_pedidos if total_pedidos else 0
    entregues     = sum(1 for o in filtered if o.get('status') == 'entregue')
    pendentes     = total_pedidos - entregues
    unidades      = sum(order_qty(o) for o in filtered)

    vol_map, val_map = {}, {}
    for o in filtered:
        items = o.get('items') or [{'produto_name': o.get('produto_name','?'),
                                     'quantidade':   o.get('quantidade', 0),
                                     'valor_total':  o.get('valor_total', 0)}]
        for it in items:
            nm  = it.get('produto_name', '?')
            qty = int(it.get('quantidade') or 0)
            vt  = float(it.get('valor_total') or 0)
            vol_map[nm] = vol_map.get(nm, 0)   + qty
            val_map[nm] = val_map.get(nm, 0.0) + vt

    por_volume = sorted(vol_map.items(), key=lambda x: x[1], reverse=True)[:10]
    por_valor  = sorted(val_map.items(), key=lambda x: x[1], reverse=True)[:10]

    dia_map = {}
    for o in filtered:
        d = order_date(o)
        if not d: continue
        v = float(o.get('valor_total') or 0)
        if d not in dia_map:
            dia_map[d] = {'dia': d, 'valor': 0.0, 'pedidos': 0}
        dia_map[d]['valor']   += v
        dia_map[d]['pedidos'] += 1
    por_dia = sorted(dia_map.values(), key=lambda x: x['dia'])

    return jsonify({
        'kpis': {
            'total_pedidos': total_pedidos,
            'valor_total':   round(valor_total, 2),
            'ticket_medio':  round(ticket_medio, 2),
            'unidades':      unidades,
            'entregues':     entregues,
            'pendentes':     pendentes,
        },
        'por_volume': [{'nome': k, 'qty': v}           for k, v in por_volume],
        'por_valor':  [{'nome': k, 'valor': round(v,2)} for k, v in por_valor],
        'por_dia':    por_dia,
    })

# ── Página Admin ───────────────────────────────────────────────────────────────
@app.get('/admin')
@require_admin
def admin():
    total = len(load_orders())
    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Admin — Lojinha Genomma</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#F3EDF9;min-height:100vh}}
header{{background:linear-gradient(135deg,#4A1B7A,#1A5FB4);padding:22px 32px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px}}
header h1{{color:white;font-size:1.4rem;font-weight:800}}
header p{{color:rgba(255,255,255,.75);font-size:.88rem}}
.badge{{background:rgba(255,255,255,.2);color:white;padding:5px 14px;border-radius:20px;font-size:.83rem;font-weight:700}}
.container{{max-width:1250px;margin:0 auto;padding:24px 18px}}
/* stats */
.stats{{display:flex;gap:13px;flex-wrap:wrap;margin-bottom:20px}}
.stat{{background:white;border-radius:13px;padding:15px 19px;flex:1;min-width:120px;box-shadow:0 2px 10px rgba(74,27,122,.09);border-left:4px solid #4A1B7A}}
.stat.green{{border-color:#27AE60}} .stat.orange{{border-color:#E67E22}}
.stat .n{{font-size:1.7rem;font-weight:800;color:#4A1B7A}}
.stat.green .n{{color:#27AE60}} .stat.orange .n{{color:#E67E22}}
.stat .l{{font-size:.75rem;color:#999;margin-top:1px}}
/* filtros */
.filters{{background:white;border-radius:13px;padding:16px 20px;margin-bottom:18px;box-shadow:0 2px 10px rgba(74,27,122,.09);display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end}}
.fgroup label{{font-size:.72rem;font-weight:700;color:#7B3FAD;text-transform:uppercase;letter-spacing:.4px;display:block;margin-bottom:4px}}
.fgroup input,.fgroup select{{border:2px solid rgba(74,27,122,.15);border-radius:8px;padding:7px 11px;font-size:.85rem;color:#333;outline:none;background:white;min-width:130px}}
.fgroup input:focus,.fgroup select:focus{{border-color:#7B3FAD}}
.btn{{border:none;cursor:pointer;border-radius:8px;padding:8px 18px;font-size:.85rem;font-weight:700;transition:.18s}}
.btn-primary{{background:linear-gradient(135deg,#4A1B7A,#1A5FB4);color:white}}
.btn-primary:hover{{opacity:.88}}
.btn-ghost{{background:white;color:#7B3FAD;border:2px solid rgba(74,27,122,.2)}}
.btn-ghost:hover{{background:#F3EDF9}}
/* card */
.card{{background:white;border-radius:15px;box-shadow:0 2px 14px rgba(74,27,122,.09);overflow:hidden}}
.card-head{{padding:15px 20px;border-bottom:1px solid #F0E8FB;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}}
.card-head h2{{font-size:.92rem;font-weight:700;color:#4A1B7A}}
.count-lbl{{font-size:.78rem;color:#aaa}}
/* tabela */
table{{width:100%;border-collapse:collapse}}
thead th{{background:#F8F3FF;padding:9px 12px;text-align:left;font-size:.7rem;font-weight:700;color:#7B3FAD;text-transform:uppercase;letter-spacing:.5px;white-space:nowrap}}
tbody tr{{border-bottom:1px solid #FAF5FF;transition:background .13s}}
tbody tr:hover{{background:#FAF5FF}}
tbody tr.entregue{{background:#F0FFF4}}
tbody tr.entregue:hover{{background:#E8FFF0}}
td{{padding:10px 12px;font-size:.86rem;vertical-align:middle}}
/* badges */
.tg{{display:inline-block;padding:3px 10px;border-radius:20px;font-size:.72rem;font-weight:700;white-space:nowrap}}
.tg-g{{background:#E8F5E9;color:#2E7D32}}
.tg-t{{background:#E3F2FD;color:#1565C0}}
.tg-p{{background:#FFF8E1;color:#E65100}}
.tg-e{{background:#E8F5E9;color:#1B5E20}}
/* botão entrega */
.btn-del{{border:none;cursor:pointer;border-radius:7px;padding:5px 12px;font-size:.76rem;font-weight:700;transition:.18s;white-space:nowrap}}
.btn-del.pend{{background:#4A1B7A;color:white}} .btn-del.pend:hover{{background:#6B2FA0}}
.btn-del.done{{background:#E8F5E9;color:#2E7D32;border:1.5px solid #A5D6A7}}
.btn-del.done:hover{{background:#FFEBEE;color:#c62828;border-color:#EF9A9A}}
.btn-exc{{border:none;cursor:pointer;border-radius:7px;padding:5px 10px;font-size:.76rem;font-weight:700;transition:.18s;background:#FFF0F0;color:#c62828;border:1.5px solid #FFCDD2;white-space:nowrap}}
.btn-exc:hover{{background:#FFEBEE;border-color:#EF9A9A}}
/* empty */
.empty{{text-align:center;padding:50px;color:#ccc}}
.empty span{{font-size:2.5rem;display:block;margin-bottom:10px}}
.rbtn{{background:#4A1B7A;color:white;border:none;padding:7px 15px;border-radius:8px;cursor:pointer;font-size:.82rem;font-weight:600}}
.rbtn:hover{{background:#6B2FA0}}
@media(max-width:750px){{
  thead{{display:none}} .stat .n{{font-size:1.3rem}}
  tbody tr{{display:block;padding:12px;border-bottom:2px solid #F0E8FB}}
  td{{display:block;padding:2px 0}}
}}
/* ── Modal relatório ── */
.rel-overlay{{position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:1000;display:none;align-items:flex-start;justify-content:center;padding:24px 12px;overflow-y:auto}}
.rel-overlay.open{{display:flex}}
.rel-panel{{background:white;border-radius:18px;width:100%;max-width:920px;box-shadow:0 12px 48px rgba(0,0,0,.3);animation:slideDown .22s ease}}
@keyframes slideDown{{from{{opacity:0;transform:translateY(-24px)}}to{{opacity:1;transform:translateY(0)}}}}
.rel-hdr{{background:linear-gradient(135deg,#4A1B7A,#1A5FB4);border-radius:18px 18px 0 0;padding:20px 26px;display:flex;align-items:center;justify-content:space-between}}
.rel-hdr h2{{color:white;font-size:1.2rem;font-weight:800;margin:0}}
.rel-close{{background:rgba(255,255,255,.2);border:none;color:white;font-size:1.2rem;width:36px;height:36px;border-radius:50%;cursor:pointer;display:flex;align-items:center;justify-content:center;transition:.18s}}
.rel-close:hover{{background:rgba(255,255,255,.35)}}
.rel-body{{padding:22px 26px}}
.rel-filters{{display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end;margin-bottom:22px;background:#F8F3FF;border-radius:12px;padding:16px}}
.rel-filters label{{font-size:.72rem;font-weight:700;color:#7B3FAD;text-transform:uppercase;letter-spacing:.4px;display:block;margin-bottom:4px}}
.rel-filters input{{border:2px solid rgba(74,27,122,.18);border-radius:8px;padding:7px 11px;font-size:.85rem;color:#333;outline:none}}
.rel-filters input:focus{{border-color:#7B3FAD}}
.rel-kpis{{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:12px;margin-bottom:24px}}
.rel-kpi{{background:#F8F3FF;border-radius:12px;padding:14px 16px;border-left:4px solid #7B3FAD;text-align:center}}
.rel-kpi.green{{border-color:#27AE60;background:#F0FFF4}}
.rel-kpi.blue{{border-color:#1A5FB4;background:#EEF4FF}}
.rel-kpi.orange{{border-color:#E67E22;background:#FFF8F0}}
.rel-kpi .kv{{font-size:1.5rem;font-weight:800;color:#4A1B7A}}
.rel-kpi.green .kv{{color:#27AE60}}
.rel-kpi.blue .kv{{color:#1A5FB4}}
.rel-kpi.orange .kv{{color:#E67E22}}
.rel-kpi .kl{{font-size:.72rem;color:#999;margin-top:2px}}
.rel-charts{{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-bottom:20px}}
@media(max-width:640px){{.rel-charts{{grid-template-columns:1fr}}}}
.rel-chart-box{{background:#F8F3FF;border-radius:12px;padding:16px}}
.rel-chart-box h3{{font-size:.82rem;font-weight:700;color:#7B3FAD;margin-bottom:12px;text-transform:uppercase;letter-spacing:.4px}}
.rel-day-box{{background:#F8F3FF;border-radius:12px;padding:16px;margin-bottom:4px}}
.rel-day-box h3{{font-size:.82rem;font-weight:700;color:#7B3FAD;margin-bottom:12px;text-transform:uppercase;letter-spacing:.4px}}
.rel-empty{{text-align:center;padding:40px;color:#bbb;font-size:.95rem}}
</style>
</head>
<body>
<header>
  <div><h1>🛍️ Painel de Pedidos</h1><p>Lojinha Interna — Genomma Lab</p></div>
  <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap">
    <span class="badge" id="hdr-count">📦 {total} pedido(s)</span>
    <button class="rbtn" onclick="loadOrders()">↻ Atualizar</button>
    <button class="rbtn" onclick="openRelatorio()" style="background:rgba(255,183,0,.85);color:#2d1a00;">📈 Relatório</button>
    <a href="/inventario" class="rbtn" style="text-decoration:none;background:rgba(39,174,96,.35)">📊 Inventário</a>
    <a href="/" class="rbtn" style="text-decoration:none">🛍️ Lojinha</a>
  </div>
</header>

<div class="container">
  <div class="stats">
    <div class="stat">       <div class="n" id="st-total">—</div>   <div class="l">Total de pedidos</div></div>
    <div class="stat">       <div class="n" id="st-genomma">—</div> <div class="l">Genomma</div></div>
    <div class="stat">       <div class="n" id="st-terc">—</div>    <div class="l">Terceirizados</div></div>
    <div class="stat">       <div class="n" id="st-units">—</div>   <div class="l">Unidades pedidas</div></div>
    <div class="stat green"> <div class="n" id="st-done">—</div>    <div class="l">✅ Entregues</div></div>
    <div class="stat orange"><div class="n" id="st-pend">—</div>    <div class="l">⏳ Pendentes</div></div>
    <div class="stat" style="border-color:#1A5FB4"><div class="n" style="color:#1A5FB4;font-size:1.2rem" id="st-valor">—</div><div class="l">💰 Valor total</div></div>
  </div>

  <div class="filters">
    <div class="fgroup"><label>De</label><input type="date" id="f-de"></div>
    <div class="fgroup"><label>Até</label><input type="date" id="f-ate"></div>
    <div class="fgroup">
      <label>Status</label>
      <select id="f-status">
        <option value="">Todos</option>
        <option value="pendente">⏳ Pendente</option>
        <option value="entregue">✅ Entregue</option>
      </select>
    </div>
    <div class="fgroup">
      <label>Tipo</label>
      <select id="f-tipo">
        <option value="">Todos</option>
        <option value="genomma">🏢 Genomma</option>
        <option value="terceirizado">🤝 Terceirizado</option>
      </select>
    </div>
    <div class="fgroup"><label>&nbsp;</label><button class="btn btn-primary" onclick="applyFilter()">🔍 Filtrar</button></div>
    <div class="fgroup"><label>&nbsp;</label><button class="btn btn-ghost"   onclick="clearFilter()">✕ Limpar</button></div>
  </div>

  <div class="card">
    <div class="card-head">
      <h2>📋 Pedidos (mais recente primeiro)</h2>
      <span class="count-lbl" id="tbl-count"></span>
    </div>
    <div id="tbl-wrap" style="overflow-x:auto"></div>
  </div>
</div>

<!-- ── Modal Relatório ─────────────────────────────────────────────────── -->
<div class="rel-overlay" id="rel-overlay" onclick="overlayClick(event)">
  <div class="rel-panel">
    <div class="rel-hdr">
      <h2>📈 Relatório de Vendas</h2>
      <button class="rel-close" onclick="closeRelatorio()">✕</button>
    </div>
    <div class="rel-body">
      <div class="rel-filters">
        <div><label>De</label><input type="date" id="rel-de"></div>
        <div><label>Até</label><input type="date" id="rel-ate"></div>
        <div style="display:flex;align-items:flex-end">
          <button class="btn btn-primary" onclick="carregarRelatorio()" style="height:36px">📊 Gerar</button>
        </div>
        <div style="display:flex;align-items:flex-end">
          <button class="btn btn-ghost" onclick="limparRelFiltro()" style="height:36px">✕ Limpar</button>
        </div>
      </div>
      <div id="rel-content"><div class="rel-empty">Selecione um período e clique em <b>Gerar</b>.</div></div>
    </div>
  </div>
</div>

<script>
let allOrders = [];

async function loadOrders() {{
  try {{
    const r = await fetch('/api/orders');
    allOrders = await r.json();
    document.getElementById('hdr-count').textContent = '📦 ' + allOrders.length + ' pedido(s)';
    applyFilter();
  }} catch(e) {{ console.error(e); }}
}}

function orderQty(o) {{
  if (o.items) return o.items.reduce((s,i)=>s+(parseInt(i.quantidade)||0),0);
  return parseInt(o.quantidade)||0;
}}
function orderProductLabel(o) {{
  if (o.items) return o.items.map(i=>`${{i.produto_name}} (${{i.quantidade}} un.)`).join('<br>');
  return o.produto_name||'';
}}

function stats(orders) {{
  document.getElementById('st-total').textContent   = orders.length;
  document.getElementById('st-genomma').textContent = orders.filter(o=>o.tipo==='genomma').length;
  document.getElementById('st-terc').textContent    = orders.filter(o=>o.tipo==='terceirizado').length;
  document.getElementById('st-units').textContent   = orders.reduce((s,o)=>s+orderQty(o),0);
  document.getElementById('st-done').textContent    = orders.filter(o=>o.status==='entregue').length;
  document.getElementById('st-pend').textContent    = orders.filter(o=>o.status!=='entregue').length;
  const total_val = orders.reduce((s,o)=>s+(parseFloat(o.valor_total)||0),0);
  document.getElementById('st-valor').textContent   = 'R$ ' + total_val.toLocaleString('pt-BR',{{minimumFractionDigits:2,maximumFractionDigits:2}});
}}

function applyFilter() {{
  const de  = document.getElementById('f-de').value;
  const ate = document.getElementById('f-ate').value;
  const st  = document.getElementById('f-status').value;
  const tp  = document.getElementById('f-tipo').value;
  const fil = allOrders.filter(o => {{
    const p = o.data_hora ? o.data_hora.split(' ')[0].split('/') : null;
    const d = p ? p[2]+'-'+p[1]+'-'+p[0] : '';
    if (de  && d < de)  return false;
    if (ate && d > ate) return false;
    if (st  && o.status !== st) return false;
    if (tp  && o.tipo   !== tp) return false;
    return true;
  }});
  stats(fil);
  render(fil);
}}

function clearFilter() {{
  ['f-de','f-ate'].forEach(id=>document.getElementById(id).value='');
  ['f-status','f-tipo'].forEach(id=>document.getElementById(id).value='');
  stats(allOrders); render(allOrders);
}}

function render(orders) {{
  const wrap = document.getElementById('tbl-wrap');
  document.getElementById('tbl-count').textContent = orders.length + ' registro(s)';
  if (!orders.length) {{
    wrap.innerHTML = "<div class='empty'><span>📭</span>Nenhum pedido encontrado.</div>";
    return;
  }}
  let rows = '';
  orders.forEach(o => {{
    const done   = o.status === 'entregue';
    const tBadge = o.tipo==='genomma' ? "<span class='tg tg-g'>🏢 Genomma</span>" : "<span class='tg tg-t'>🤝 Terceirizado</span>";
    const sBadge = done
      ? "<span class='tg tg-e'>✅ Entregue" + (o.entregue_em ? '<br><small style=\\"font-weight:400;opacity:.75;font-size:.68rem\\">' + o.entregue_em + '</small>' : '') + "</span>"
      : "<span class='tg tg-p'>⏳ Pendente</span>";
    const cLink  = o.comprovante ? `<a href="/uploads/${{o.comprovante}}" target="_blank" style="color:#4A1B7A;font-weight:600;text-decoration:none;">📎 Ver</a>` : '—';
    const nfCell = o.nota_fiscal
      ? `<a href="/uploads/${{o.nota_fiscal}}" target="_blank" style="color:#27AE60;font-weight:600;text-decoration:none;">📄 Ver NF</a>`
      : `<label style="cursor:pointer;background:#EAF7EE;color:#27AE60;border:1px solid #27AE60;border-radius:6px;padding:3px 8px;font-size:.78rem;font-weight:600;white-space:nowrap;">
           📎 Anexar NF
           <input type="file" accept=".pdf,.jpg,.jpeg,.png" style="display:none" onchange="anexarNF('${{o.id}}',this)">
         </label>`;
    const newSt  = done ? 'pendente' : 'entregue';
    const btnLbl = done ? '↩ Desfazer' : '✅ Marcar entregue';
    const btnCls = done ? 'done' : 'pend';
    const fmtBRL = v => v != null && v > 0 ? 'R$ ' + parseFloat(v).toLocaleString('pt-BR',{{minimumFractionDigits:2,maximumFractionDigits:2}}) : '—';
    const base = `
      <td style="white-space:nowrap;color:#666;font-size:.78rem;">${{o.data_hora||''}}</td>
      <td style="font-weight:600">${{o.nome||''}}</td>
      <td style="color:#4A1B7A;font-size:.82rem;">${{o.email||''}}</td>
      <td>${{tBadge}}</td>`;
    const tail = `
      <td style="text-align:center;font-weight:700;color:#1A5FB4;">${{fmtBRL(o.valor_total)}}</td>
      <td style="text-align:center">${{cLink}}</td>
      <td style="text-align:center">${{nfCell}}</td>
      <td style="text-align:center">${{sBadge}}</td>
      <td style="text-align:center"><button class="btn-del ${{btnCls}}" onclick="toggle('${{o.id}}','${{newSt}}')">${{btnLbl}}</button></td>
      <td style="text-align:center"><button class="btn-exc" onclick="excluir('${{o.id}}')">🗑️ Excluir</button></td>`;
    let mid = '';
    if (o.items && o.items.length) {{
      const miniRows = o.items.map(it =>
        `<div style="display:flex;align-items:center;gap:6px;padding:3px 0;border-bottom:1px solid #EDE3F5;flex-wrap:wrap;">
          <span style="flex:1;min-width:120px;font-size:.8rem;color:#333;">${{it.produto_name}}</span>
          <span style="font-size:.8rem;font-weight:700;color:#27AE60;white-space:nowrap;">${{it.quantidade}} un.</span>
          <span style="font-size:.8rem;color:#888;white-space:nowrap;">${{fmtBRL(it.preco_unit)}}/un</span>
          <span style="font-size:.8rem;font-weight:700;color:#1A5FB4;white-space:nowrap;">${{fmtBRL(it.valor_total)}}</span>
        </div>`
      ).join('');
      mid = `<td colspan="3" style="padding:6px 8px;">${{miniRows}}</td>`;
    }} else {{
      mid = `
        <td style="max-width:180px;font-size:.82rem;">${{o.produto_name||''}}</td>
        <td style="text-align:center;font-weight:700;color:#27AE60;font-size:1rem;">${{o.quantidade||''}}</td>
        <td style="text-align:center;font-size:.82rem;color:#555;">${{fmtBRL(o.preco_unit)}}</td>`;
    }}
    rows += `<tr class="${{done?'entregue':''}}" id="row-${{o.id}}">${{base}}${{mid}}${{tail}}</tr>`;
  }});
  wrap.innerHTML = `<table><thead><tr>
    <th>Data/Hora</th><th>Nome</th><th>E-mail</th><th>Tipo</th>
    <th>Produto</th><th style="text-align:center">Qtd</th><th style="text-align:center">Preço Unit.</th>
    <th style="text-align:center">Valor Total</th>
    <th style="text-align:center">Comprovante</th>
    <th style="text-align:center">Nota Fiscal</th>
    <th style="text-align:center">Status</th><th style="text-align:center">Entrega</th>
    <th style="text-align:center">Excluir</th>
  </tr></thead><tbody>${{rows}}</tbody></table>`;
}}

async function toggle(id, newStatus) {{
  const btn = document.querySelector(`#row-${{id}} .btn-del`);
  if (btn) {{ btn.disabled=true; btn.textContent='...'; }}
  try {{
    const r = await fetch(`/api/orders/${{id}}/status`,{{
      method:'POST', headers:{{'Content-Type':'application/json'}},
      body: JSON.stringify({{status:newStatus}})
    }});
    if (r.ok) {{ await loadOrders(); }}
    else {{ alert('Erro ao atualizar. Tente novamente.'); if(btn) btn.disabled=false; }}
  }} catch(e) {{ alert('Erro de conexão.'); if(btn) btn.disabled=false; }}
}}

async function excluir(id) {{
  const row = document.getElementById(`row-${{id}}`);
  const prod = row ? row.querySelector('td:nth-child(5)').innerText : '';
  if (!confirm(`Excluir este pedido?\n\n${{prod}}\n\nO estoque será restaurado automaticamente.`)) return;
  try {{
    const r = await fetch(`/api/orders/${{id}}`, {{method: 'DELETE'}});
    if (r.ok) {{ await loadOrders(); }}
    else {{ alert('Erro ao excluir. Tente novamente.'); }}
  }} catch(e) {{ alert('Erro de conexão.'); }}
}}

async function anexarNF(id, input) {{
  const file = input.files[0];
  if (!file) return;
  const label = input.parentElement;
  label.textContent = '⏳ Enviando...';
  const fd = new FormData();
  fd.append('nota_fiscal', file);
  try {{
    const r = await fetch(`/api/orders/${{id}}/nota_fiscal`, {{method:'POST', body:fd}});
    if (r.ok) {{
      await loadOrders();
    }} else {{
      const d = await r.json();
      alert('Erro: ' + (d.error||'Falha no upload.'));
      label.innerHTML = '📎 Anexar NF <input type="file" accept=".pdf,.jpg,.jpeg,.png" style="display:none" onchange="anexarNF(\\'' + id + '\\',this)">';
    }}
  }} catch(e) {{
    alert('Erro de conexão.');
    label.innerHTML = '📎 Anexar NF <input type="file" accept=".pdf,.jpg,.jpeg,.png" style="display:none" onchange="anexarNF(\\'' + id + '\\',this)">';
  }}
}}

// ── Relatório ─────────────────────────────────────────────────────────────────
let _chartVol = null, _chartVal = null, _chartDia = null;

function openRelatorio() {{
  // default: last 30 days
  const today = new Date();
  const from  = new Date(today); from.setDate(today.getDate()-30);
  const fmt = d => d.toISOString().slice(0,10);
  if (!document.getElementById('rel-de').value)  document.getElementById('rel-de').value  = fmt(from);
  if (!document.getElementById('rel-ate').value) document.getElementById('rel-ate').value = fmt(today);
  document.getElementById('rel-overlay').classList.add('open');
  carregarRelatorio();
}}

function closeRelatorio() {{
  document.getElementById('rel-overlay').classList.remove('open');
}}

function overlayClick(e) {{
  if (e.target === document.getElementById('rel-overlay')) closeRelatorio();
}}

function limparRelFiltro() {{
  document.getElementById('rel-de').value  = '';
  document.getElementById('rel-ate').value = '';
  document.getElementById('rel-content').innerHTML = '<div class="rel-empty">Selecione um período e clique em <b>Gerar</b>.</div>';
  if (_chartVol) {{ _chartVol.destroy(); _chartVol = null; }}
  if (_chartVal) {{ _chartVal.destroy(); _chartVal = null; }}
  if (_chartDia) {{ _chartDia.destroy(); _chartDia = null; }}
}}

function fmtBRLRel(v) {{
  return 'R$ ' + parseFloat(v).toLocaleString('pt-BR',{{minimumFractionDigits:2,maximumFractionDigits:2}});
}}

async function carregarRelatorio() {{
  const de  = document.getElementById('rel-de').value;
  const ate = document.getElementById('rel-ate').value;
  const box = document.getElementById('rel-content');
  box.innerHTML = '<div class="rel-empty">⏳ Carregando...</div>';
  if (_chartVol) {{ _chartVol.destroy(); _chartVol = null; }}
  if (_chartVal) {{ _chartVal.destroy(); _chartVal = null; }}
  if (_chartDia) {{ _chartDia.destroy(); _chartDia = null; }}
  try {{
    const qs  = new URLSearchParams();
    if (de)  qs.set('de',  de);
    if (ate) qs.set('ate', ate);
    const r   = await fetch('/api/relatorio?' + qs.toString());
    const data = await r.json();
    renderRelatorio(data);
  }} catch(e) {{
    box.innerHTML = '<div class="rel-empty">❌ Erro ao carregar relatório.</div>';
  }}
}}

function renderRelatorio(data) {{
  const k = data.kpis;
  const box = document.getElementById('rel-content');

  if (k.total_pedidos === 0) {{
    box.innerHTML = '<div class="rel-empty">📭 Nenhum pedido no período selecionado.</div>';
    return;
  }}

  const entregPct = k.total_pedidos ? Math.round(k.entregues/k.total_pedidos*100) : 0;

  box.innerHTML = `
    <div class="rel-kpis">
      <div class="rel-kpi blue">
        <div class="kv">${{k.total_pedidos}}</div>
        <div class="kl">Total de Pedidos</div>
      </div>
      <div class="rel-kpi blue">
        <div class="kv" style="font-size:1.1rem">${{fmtBRLRel(k.valor_total)}}</div>
        <div class="kl">💰 Valor Total</div>
      </div>
      <div class="rel-kpi">
        <div class="kv" style="font-size:1.1rem">${{fmtBRLRel(k.ticket_medio)}}</div>
        <div class="kl">🎫 Ticket Médio</div>
      </div>
      <div class="rel-kpi orange">
        <div class="kv">${{k.unidades}}</div>
        <div class="kl">📦 Unidades Vendidas</div>
      </div>
      <div class="rel-kpi green">
        <div class="kv">${{k.entregues}}</div>
        <div class="kl">✅ Entregues (${{entregPct}}%)</div>
      </div>
      <div class="rel-kpi" style="border-color:#E67E22;background:#FFF8F0">
        <div class="kv" style="color:#E67E22">${{k.pendentes}}</div>
        <div class="kl">⏳ Pendentes</div>
      </div>
    </div>
    <div class="rel-charts">
      <div class="rel-chart-box">
        <h3>🏆 Top Produtos — Volume (un.)</h3>
        <canvas id="chart-vol" height="220"></canvas>
      </div>
      <div class="rel-chart-box">
        <h3>💰 Top Produtos — Valor (R$)</h3>
        <canvas id="chart-val" height="220"></canvas>
      </div>
    </div>
    ${{data.por_dia.length > 1 ? '<div class="rel-day-box"><h3>📅 Vendas por Dia (R$)</h3><canvas id="chart-dia" height="140"></canvas></div>' : ''}}
  `;

  const PURP  = 'rgba(74,27,122,';
  const BLUE  = 'rgba(26,95,180,';
  const GREEN = 'rgba(39,174,96,';

  const volLabels = data.por_volume.map(x => x.nome.length>22 ? x.nome.slice(0,20)+'…' : x.nome);
  const volData   = data.por_volume.map(x => x.qty);
  _chartVol = new Chart(document.getElementById('chart-vol'), {{
    type: 'bar',
    data: {{
      labels: volLabels,
      datasets: [{{ label: 'Unidades', data: volData,
        backgroundColor: volData.map((_,i) => PURP+(i===0?'1)':i===1?'.8)':'.6)')),
        borderRadius: 6, borderSkipped: false }}]
    }},
    options: {{
      indexAxis: 'y', responsive: true, plugins: {{ legend: {{ display:false }},
        tooltip: {{ callbacks: {{ label: ctx => ' '+ctx.raw+' un.' }} }} }},
      scales: {{ x: {{ grid: {{ color:'rgba(0,0,0,.05)' }}, ticks: {{ font:{{size:11}} }} }},
                 y: {{ ticks: {{ font:{{size:11}} }} }} }}
    }}
  }});

  const valLabels = data.por_valor.map(x => x.nome.length>22 ? x.nome.slice(0,20)+'…' : x.nome);
  const valData   = data.por_valor.map(x => x.valor);
  _chartVal = new Chart(document.getElementById('chart-val'), {{
    type: 'bar',
    data: {{
      labels: valLabels,
      datasets: [{{ label: 'Valor', data: valData,
        backgroundColor: valData.map((_,i) => BLUE+(i===0?'1)':i===1?'.8)':'.6)')),
        borderRadius: 6, borderSkipped: false }}]
    }},
    options: {{
      indexAxis: 'y', responsive: true, plugins: {{ legend: {{ display:false }},
        tooltip: {{ callbacks: {{ label: ctx => ' R$ '+parseFloat(ctx.raw).toLocaleString('pt-BR',{{minimumFractionDigits:2}}) }} }} }},
      scales: {{ x: {{ grid: {{ color:'rgba(0,0,0,.05)' }},
                       ticks: {{ font:{{size:11}}, callback: v => 'R$'+v.toLocaleString('pt-BR') }} }},
                 y: {{ ticks: {{ font:{{size:11}} }} }} }}
    }}
  }});

  if (data.por_dia.length > 1) {{
    const diaLabels = data.por_dia.map(x => x.dia.split('-').reverse().join('/'));
    const diaValor  = data.por_dia.map(x => x.valor);
    _chartDia = new Chart(document.getElementById('chart-dia'), {{
      type: 'line',
      data: {{
        labels: diaLabels,
        datasets: [{{ label: 'Valor (R$)', data: diaValor, fill: true,
          backgroundColor: 'rgba(39,174,96,.12)', borderColor: GREEN+'1)',
          pointBackgroundColor: GREEN+'1)', tension: 0.35, borderWidth: 2.5 }}]
      }},
      options: {{
        responsive: true, plugins: {{ legend: {{ display:false }},
          tooltip: {{ callbacks: {{ label: ctx => ' R$ '+parseFloat(ctx.raw).toLocaleString('pt-BR',{{minimumFractionDigits:2}}) }} }} }},
        scales: {{ x: {{ grid: {{ color:'rgba(0,0,0,.04)' }}, ticks: {{ font:{{size:11}}, maxRotation:45 }} }},
                   y: {{ grid: {{ color:'rgba(0,0,0,.04)' }},
                         ticks: {{ font:{{size:11}}, callback: v => 'R$'+v.toLocaleString('pt-BR') }} }} }}
      }}
    }});
  }}
}}

loadOrders();
setInterval(loadOrders, 30000);
</script>
</body></html>"""
    return Response(html, mimetype='text/html')

# ── Uploads ───────────────────────────────────────────────────────────────────
@app.get('/uploads/<filename>')
def serve_upload(filename):
    return send_from_directory(str(UPLOAD), filename)

# ── Frontend ───────────────────────────────────────────────────────────────────
@app.get('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')

# ── Notificação Teams ────────────────────────────────────────────────────────
def _send_teams_notification(order):
    """Envia um Adaptive Card para o canal do Teams via webhook (não bloqueia o pedido se falhar)."""
    if not TEAMS_WEBHOOK_URL:
        return
    try:
        tipo_label = '🏢 Genomma' if order.get('tipo') == 'genomma' else '🤝 Terceirizado(a)'

        if order.get('items'):
            itens = order['items']
            produtos_facts = [
                {'title': it.get('produto_name', '?'),
                 'value': f"{it.get('quantidade', 0)} un. · R$ {float(it.get('valor_total', 0)):.2f}"}
                for it in itens
            ]
        else:
            produtos_facts = [
                {'title': order.get('produto_name', '?'),
                 'value': f"{order.get('quantidade', 0)} un. · R$ {float(order.get('valor_total', 0)):.2f}"}
            ]

        card = {
            "type": "AdaptiveCard",
            "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
            "version": "1.4",
            "body": [
                {
                    "type": "TextBlock",
                    "text": "🛍️ Novo Pedido — Lojinha Genomma",
                    "weight": "Bolder",
                    "size": "Medium",
                    "wrap": True
                },
                {
                    "type": "FactSet",
                    "facts": [
                        {"title": "Nome:",   "value": order.get('nome', '')},
                        {"title": "E-mail:", "value": order.get('email', '')},
                        {"title": "Vínculo:","value": tipo_label},
                        {"title": "Data:",   "value": order.get('data_hora', '')},
                    ]
                },
                {
                    "type": "TextBlock",
                    "text": "📦 Produtos",
                    "weight": "Bolder",
                    "spacing": "Medium"
                },
                {
                    "type": "FactSet",
                    "facts": produtos_facts
                },
                {
                    "type": "TextBlock",
                    "text": f"💰 Valor total: R$ {float(order.get('valor_total', 0)):.2f}",
                    "weight": "Bolder",
                    "color": "Good",
                    "spacing": "Medium"
                }
            ]
        }
        payload = {
            "type": "message",
            "attachments": [
                {
                    "contentType": "application/vnd.microsoft.card.adaptive",
                    "content": card
                }
            ]
        }
        data = json.dumps(payload).encode('utf-8')
        req  = urllib.request.Request(
            TEAMS_WEBHOOK_URL, data=data,
            headers={'Content-Type': 'application/json'}, method='POST'
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            if resp.status not in (200, 202):
                log.warning(f'⚠️  Teams webhook respondeu HTTP {resp.status}')
            else:
                log.info('📲 Notificação Teams enviada.')
    except Exception as e:
        log.warning(f'⚠️  Falha ao enviar notificação Teams: {e}')

# ── Email ─────────────────────────────────────────────────────────────────────
def _send_email(nome, email, tipo, product, qty, attach_path, attach_name):
    if not SMTP_USER or not SMTP_PASS: return
    tipo_label = '🏢 Genomma' if tipo == 'genomma' else '🤝 Terceirizado(a)'
    dh = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    msg = MIMEMultipart('alternative')
    msg['Subject'] = f'🛍️ Novo Pedido — {nome}'
    msg['From']    = f'"Lojinha Genomma" <{SMTP_USER}>'
    msg['To']      = DEST_EMAIL
    msg['Reply-To']= email
    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;background:#f5f5f5;padding:20px;">
<div style="max-width:600px;margin:auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.1);">
  <div style="background:linear-gradient(135deg,#4A1B7A,#1A5FB4);padding:28px;text-align:center;">
    <h1 style="color:white;margin:0;font-size:22px;">🛍️ Novo Pedido Recebido</h1>
    <p style="color:rgba(255,255,255,.8);margin:6px 0 0;">{dh}</p>
  </div>
  <div style="padding:28px;">
    <h2 style="color:#4A1B7A;border-bottom:2px solid #f0e6f6;padding-bottom:10px;">👤 Comprador</h2>
    <table style="width:100%;border-collapse:collapse;">
      <tr><td style="padding:7px 0;color:#666;width:140px;"><b>Nome:</b></td><td>{nome}</td></tr>
      <tr><td style="padding:7px 0;color:#666;"><b>E-mail:</b></td><td>{email}</td></tr>
      <tr><td style="padding:7px 0;color:#666;"><b>Vínculo:</b></td><td>{tipo_label}</td></tr>
    </table>
    <h2 style="color:#4A1B7A;border-bottom:2px solid #f0e6f6;padding-bottom:10px;margin-top:22px;">📦 Produto</h2>
    <table style="width:100%;border-collapse:collapse;">
      <tr><td style="padding:7px 0;color:#666;width:140px;"><b>Produto:</b></td><td>{product["name"]}</td></tr>
      <tr><td style="padding:7px 0;color:#666;"><b>Código:</b></td><td>{product["code"]}</td></tr>
      <tr><td style="padding:7px 0;color:#666;"><b>Quantidade:</b></td>
          <td style="font-size:18px;font-weight:bold;color:#27AE60;">{qty} un.</td></tr>
    </table>
    <div style="background:#f8f3ff;border-left:4px solid #4A1B7A;padding:13px;border-radius:4px;margin-top:22px;">
      <p style="margin:0;color:#4A1B7A;"><b>💡 Ação:</b> Separar {qty} un. de <em>{product["name"]}</em> → enviar para {email}</p>
    </div>
  </div>
  <div style="background:#f5f5f5;padding:14px;text-align:center;">
    <p style="color:#aaa;font-size:11px;margin:0;">Lojinha Interna Genomma Lab · {dh}</p>
  </div>
</div></body></html>"""
    msg.attach(MIMEText(html, 'html', 'utf-8'))
    if attach_path and Path(attach_path).exists():
        with open(str(attach_path),'rb') as f:
            part = MIMEBase('application','octet-stream')
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition','attachment',filename=attach_name or 'comprovante')
        full = MIMEMultipart('mixed')
        for k in ('Subject','From','To','Reply-To'): full[k] = msg[k]
        full.attach(msg); full.attach(part)
        send_msg = full
    else:
        send_msg = msg
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
        s.ehlo(); s.starttls(); s.login(SMTP_USER, SMTP_PASS)
        s.sendmail(SMTP_USER, [DEST_EMAIL], send_msg.as_string())

# ── Inicialização do estoque (compatível com gunicorn) ─────────────────────────
init_stock()

# ── Start ─────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    log.info(f'\n🚀 Lojinha            → http://localhost:{PORT}')
    log.info(f'🔧 Painel de pedidos  → http://localhost:{PORT}/admin')
    log.info(f'📧 Destino do email   → {DEST_EMAIL}')
    log.info(f'📬 SMTP               → {"configurado ("+SMTP_USER+")" if SMTP_USER else "não configurado — pedidos salvos em data/orders.json"}\n')
    app.run(host='0.0.0.0', port=PORT, debug=False)
